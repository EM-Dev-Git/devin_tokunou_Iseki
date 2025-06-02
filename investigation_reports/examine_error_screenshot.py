import pandas as pd
import openpyxl

print("=== ERROR SCREENSHOT FILE ANALYSIS ===")
print()

try:
    df = pd.read_excel('/home/ubuntu/attachments/6d1b4af2-6c03-47af-90e8-53037be92735/.xlsx', sheet_name=None)
    print("1. PANDAS ANALYSIS:")
    for sheet_name, sheet_df in df.items():
        print(f"   Sheet '{sheet_name}': {sheet_df.shape} rows/cols")
        if len(sheet_df) > 0:
            print(f"   Columns: {list(sheet_df.columns)}")
            print(f"   Sample data:")
            print(sheet_df.head())
        else:
            print("   Empty sheet")
    print()
except Exception as e:
    print(f"Pandas error: {e}")

try:
    wb = openpyxl.load_workbook('/home/ubuntu/attachments/6d1b4af2-6c03-47af-90e8-53037be92735/.xlsx')
    print("2. OPENPYXL ANALYSIS:")
    print(f"   Worksheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   Sheet '{sheet_name}':")
        print(f"     Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        if ws.max_row > 0 and ws.max_column > 0:
            print("     Cell values (first 10 rows, 5 cols):")
            for row in range(1, min(11, ws.max_row + 1)):
                row_values = []
                for col in range(1, min(6, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_values.append(str(cell_value) if cell_value is not None else "")
                print(f"       Row {row}: {row_values}")
        print()
        
except Exception as e:
    print(f"Openpyxl error: {e}")

import os
try:
    file_path = '/home/ubuntu/attachments/6d1b4af2-6c03-47af-90e8-53037be92735/.xlsx'
    file_size = os.path.getsize(file_path)
    print(f"3. FILE PROPERTIES:")
    print(f"   File size: {file_size:,} bytes")
    print(f"   File exists: {os.path.exists(file_path)}")
except Exception as e:
    print(f"File properties error: {e}")
