import openpyxl
import os
import shutil
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter

def copy_and_modify_flow_document():
    """Copy and modify the flow document according to requirements."""
    # Source and destination paths
    source_file = "ドキュメント/010.システム全般/請求処理フロー.xlsx"
    dest_file = "改修後ドキュメント3/010.システム全般/請求処理フロー.xlsx"
    
    # First copy the file
    shutil.copy2(source_file, dest_file)
    print(f"Copied {source_file} to {dest_file}")
    
    # Then modify it
    wb = openpyxl.load_workbook(dest_file)
    
    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Add "(改修版)" to the title
        if sheet['A1'].value:
            sheet['A1'].value = sheet['A1'].value + " (改修版)"
        
        # Look for rows that might contain flow information about invoice master registration
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and "請求マスター登録" in cell_value:
                print(f"Found '請求マスター登録' at row {row}")
                
                # Add automatic calculation step
                insert_row = row + 1
                sheet.insert_rows(insert_row)
                sheet.cell(row=insert_row, column=1).value = "  ↓"
                sheet.cell(row=insert_row, column=2).value = "数量×単価の自動計算"
                sheet.cell(row=insert_row, column=3).value = "数量または単価の入力/変更後に金額を自動計算"
                
                # Add input validation step
                insert_row = row + 2
                sheet.insert_rows(insert_row)
                sheet.cell(row=insert_row, column=1).value = "  ↓"
                sheet.cell(row=insert_row, column=2).value = "必須項目チェック"
                sheet.cell(row=insert_row, column=3).value = "単価、数量、単位の入力チェック"
                
                # Add unit input step
                insert_row = row + 3
                sheet.insert_rows(insert_row)
                sheet.cell(row=insert_row, column=1).value = "  ↓"
                sheet.cell(row=insert_row, column=2).value = "単位入力"
                sheet.cell(row=insert_row, column=3).value = "数量の単位を入力（必須項目）"
                break
                
        # Look for rows that might contain invoice output flow information
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and "請求書出力" in cell_value:
                print(f"Found '請求書出力' at row {row}")
                
                # Add unit information retrieval step
                insert_row = row + 1
                sheet.insert_rows(insert_row)
                sheet.cell(row=insert_row, column=1).value = "  ↓"
                sheet.cell(row=insert_row, column=2).value = "単位情報の取得"
                sheet.cell(row=insert_row, column=3).value = "請求マスターから単位情報を取得して出力"
                break
    
    # Save the workbook
    wb.save(dest_file)
    print(f"Modified flow document: {dest_file}")

if __name__ == "__main__":
    copy_and_modify_flow_document()
    print("Flow document copied and modified successfully")
